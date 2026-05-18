"""
EchoPose — Financial Documents Generator
Produces: Financial Model (XLSX) + Cap Table (XLSX) + Invoice Template (XLSX)
Run: python scripts/generate_financial.py
Output: docs/financial/
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import os

OUT = 'docs/financial'
os.makedirs(OUT, exist_ok=True)

OWNER      = 'Muhammed Shazin Sadhik Kunhi Parambath'
TRADING_AS = 'EchoPose'
EMAIL      = 'shazin2889@gmail.com'

# ── Colour fills ──────────────────────────────────────────────────────────────
def fill(hex_): return PatternFill('solid', fgColor=hex_)
def font(bold=False, color='000000', size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name='Calibri')
def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bottom():
    tb = Side(style='medium')
    tn = Side(style='thin')
    return Border(left=tn, right=tn, top=tn, bottom=tb)

F_NAVY   = '0D1B2A'
F_BLUE   = '1B6CA8'
F_TEAL   = '008789'
F_GREEN  = '1A7348'
F_RED    = 'C02020'
F_ORANGE = 'D96B00'
F_GOLD   = 'B8860B'
F_LGREY  = 'F2F4F7'
F_MGREY  = 'CCD1DA'
F_WHITE  = 'FFFFFF'
F_DKGREY = '3A3F4A'

def hdr_cell(ws, row, col, text, bg=F_NAVY, fg=F_WHITE, sz=11, bold=True, wide=False):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg); c.font = font(bold=bold, color=fg, size=sz)
    c.alignment = align('center', 'center', wrap=True)
    c.border = border()
    return c

def data_cell(ws, row, col, value, fmt=None, bg=F_WHITE, bold=False, color='000000', halign='right'):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg); c.font = font(bold=bold, color=color)
    c.alignment = align(halign, 'center')
    c.border = border()
    if fmt: c.number_format = fmt
    return c

def label_cell(ws, row, col, text, bg=F_LGREY, bold=False, indent=0):
    c = ws.cell(row=row, column=col, value=(' '*indent)+text)
    c.fill = fill(bg); c.font = font(bold=bold, color=F_DKGREY)
    c.alignment = align('left', 'center', wrap=True)
    c.border = border()
    return c

def section_row(ws, row, ncols, text, bg=F_BLUE):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg); c.font = font(bold=True, color=F_WHITE, size=11)
    c.alignment = align('left', 'center')
    ws.row_dimensions[row].height = 20

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell='B2'):
    ws.freeze_panes = cell


# ══════════════════════════════════════════════════════════════════════════════
#  FINANCIAL MODEL
# ══════════════════════════════════════════════════════════════════════════════
print('Building Financial Model...')
wb = Workbook()

# ── Cover Sheet ───────────────────────────────────────────────────────────────
ws = wb.active; ws.title = 'Cover'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 60
ws.row_dimensions[1].height = 10

for r in range(2, 22):
    ws.row_dimensions[r].height = 28

cover_data = [
    (3,  'B', TRADING_AS,                         font(bold=True, color=F_TEAL, size=28), fill(F_NAVY)),
    (4,  'B', 'Financial Model — 3-Year Projection', font(bold=True, color=F_WHITE, size=16), fill(F_NAVY)),
    (5,  'B', '',                                   font(), fill(F_NAVY)),
    (6,  'B', f'Prepared by: {OWNER}',              font(bold=False, color='B0D0FF', size=11), fill(F_NAVY)),
    (7,  'B', f'Email: {EMAIL}',                    font(color='B0D0FF', size=10), fill(F_NAVY)),
    (8,  'B', 'Status: Pre-Company / Sole Trader',  font(color='D96B00', size=10, bold=True), fill(F_NAVY)),
    (9,  'B', 'Version: 1.0  |  April 2026',        font(color=F_MGREY, size=9), fill(F_NAVY)),
    (10, 'B', 'CONFIDENTIAL — COMMERCIAL IN CONFIDENCE', font(bold=True, color=F_MGREY, size=9), fill(F_NAVY)),
]
for row, col, val, fnt, fll in cover_data:
    c = ws[f'{col}{row}']; c.value = val; c.font = fnt; c.fill = fll
    c.alignment = align('left', 'center')

ws['B12'] = 'SHEETS IN THIS WORKBOOK'
ws['B12'].font = font(bold=True, color=F_NAVY, size=12)
for i, (sheet, desc) in enumerate([
    ('Revenue Model', 'Monthly MRR/ARR by tier — 3 years (36 months)'),
    ('Expenses',      'Fixed and variable cost breakdown — monthly'),
    ('P&L',           'Revenue minus expenses — gross profit and EBITDA'),
    ('Cash Flow',     'Cash in vs cash out — cumulative position'),
    ('Break-Even',    'How many customers needed to cover costs'),
    ('Assumptions',   'All editable input parameters in one place'),
], 1):
    ws.cell(row=13+i, column=2, value=f'{i}.  {sheet}').font = font(bold=True, color=F_BLUE)
    ws.cell(row=13+i, column=3, value=desc).font = font(color=F_DKGREY)
    ws.column_dimensions['C'].width = 55

# ── Assumptions Sheet ─────────────────────────────────────────────────────────
wa = wb.create_sheet('Assumptions')
wa.sheet_view.showGridLines = False
set_col_widths(wa, [3, 35, 18, 18, 45])
freeze(wa, 'C3')

wa.merge_cells('B1:E1')
c = wa['B1']; c.value = f'{TRADING_AS} — Model Assumptions (Edit These)'
c.font = font(bold=True, color=F_WHITE, size=14); c.fill = fill(F_NAVY)
c.alignment = align('center','center')
wa.row_dimensions[1].height = 30

for col, hdr in enumerate(['Parameter', 'Value', 'Unit', 'Notes'], 2):
    hdr_cell(wa, 2, col, hdr, F_BLUE)
wa.row_dimensions[2].height = 18

assumptions = [
    ('PRICING', None, None, None),
    ('Home tier price', 49, '£/month', 'EchoPose Home — families'),
    ('Care tier price', 199, '£/month', 'EchoPose Care — care agencies'),
    ('Pro tier price', 499, '£/month', 'EchoPose Pro — care homes'),
    ('Hardware kit price', 299, '£ one-time', 'ESP32 nodes + router + setup'),
    ('On-premise license (avg)', 5000, '£ one-time', 'Average across security/enterprise deals'),
    ('Annual support contract', 600, '£/year', 'For on-premise customers'),
    ('', None, None, None),
    ('GROWTH — YEAR 1 (months 1–12)', None, None, None),
    ('New Home customers/month', 2.5, 'avg/month', 'Starting at 1, reaching 4 by month 12'),
    ('New Care customers/month', 1.25, 'avg/month', 'Starting at 0.5, reaching 2 by month 12'),
    ('New Pro customers/month', 0.4, 'avg/month', 'Starting at 0, reaching 0.5 by month 12'),
    ('', None, None, None),
    ('GROWTH — YEAR 2 (months 13–24)', None, None, None),
    ('New Home customers/month', 8, 'avg/month', 'Google Ads + referrals kicking in'),
    ('New Care customers/month', 5, 'avg/month', 'LinkedIn outreach + pilot conversions'),
    ('New Pro customers/month', 1.5, 'avg/month', 'Care home sector builds slowly'),
    ('', None, None, None),
    ('GROWTH — YEAR 3 (months 25–36)', None, None, None),
    ('New Home customers/month', 18, 'avg/month', 'Channel partner + word of mouth'),
    ('New Care customers/month', 12, 'avg/month', 'Care agency network effects'),
    ('New Pro customers/month', 4, 'avg/month', 'Enterprise pipeline maturing'),
    ('', None, None, None),
    ('CHURN', None, None, None),
    ('Home monthly churn rate', 0.05, '% of active', '5% per month = 20 month avg lifetime'),
    ('Care monthly churn rate', 0.03, '% of active', '3% per month = 33 month avg lifetime'),
    ('Pro monthly churn rate', 0.02, '% of active', '2% per month = 50 month avg lifetime'),
    ('', None, None, None),
    ('COSTS', None, None, None),
    ('Cloud hosting (base)', 150, '£/month', 'DigitalOcean — scales with customer count'),
    ('Hosting per 100 customers', 40, '£/month', 'Additional hosting per 100 active customers'),
    ('Email/SMS tools (base)', 50, '£/month', 'SendGrid + Twilio base'),
    ('Software tools', 40, '£/month', 'GitHub, domains, misc SaaS'),
    ('Stripe fees', 0.014, '% of revenue', '1.4% + 20p per transaction — simplified as %'),
    ('Marketing budget Y1', 0, '£/month', 'Zero paid marketing in Year 1'),
    ('Marketing budget Y2', 1500, '£/month', 'Google Ads + occasional events'),
    ('Marketing budget Y3', 4000, '£/month', 'Scaling paid acquisition'),
    ('', None, None, None),
    ('HARDWARE / ONE-TIME', None, None, None),
    ('Hardware kits sold (% of new customers)', 0.8, 'ratio', '80% of new customers buy a kit'),
    ('Hardware COGS (% of kit price)', 0.55, 'ratio', '55% COGS on hardware (55% cost, 45% margin)'),
    ('On-premise deals per year Y1', 1.5, 'deals/year', 'Approx 1–2 deals in Year 1'),
    ('On-premise deals per year Y2', 6, 'deals/year', ''),
    ('On-premise deals per year Y3', 18, 'deals/year', ''),
]

row = 3
for item in assumptions:
    label, val, unit, note = item
    bg = F_LGREY if (row % 2 == 0) else F_WHITE
    if val is None:
        section_row(wa, row, 5, f'  {label}', F_DKGREY)
    else:
        label_cell(wa, row, 2, label, bg=bg)
        c = wa.cell(row=row, column=3, value=val)
        c.fill = fill('FFFDE7'); c.font = font(bold=True, color=F_NAVY)
        c.alignment = align('center','center')
        c.border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        if isinstance(val, float) and val < 1 and unit and '%' in unit:
            c.number_format = '0.0%'
        elif isinstance(val, (int, float)) and unit and '£' in unit:
            c.number_format = '£#,##0'
        data_cell(wa, row, 4, unit or '', bg=bg, halign='left')
        data_cell(wa, row, 5, note or '', bg=bg, halign='left')
    row += 1

wa.row_dimensions[1].height = 30
for r in range(2, row):
    h_ = wa.row_dimensions[r].height
    if h_ is None or h_ < 15:
        wa.row_dimensions[r].height = 18

# ── Revenue Model Sheet ───────────────────────────────────────────────────────
wr = wb.create_sheet('Revenue Model')
wr.sheet_view.showGridLines = False
wr.freeze_panes = 'C4'

months = [f'M{i:02d}' for i in range(1, 37)]
years  = ['Year 1'] * 12 + ['Year 2'] * 12 + ['Year 3'] * 12

wr.merge_cells('A1:AK1')
c = wr['A1']; c.value = f'{TRADING_AS} — Revenue Model (36 Months)'
c.font = font(bold=True, color=F_WHITE, size=13); c.fill = fill(F_NAVY)
c.alignment = align('center','center')
wr.row_dimensions[1].height = 28

hdr_cell(wr, 2, 1, 'Year', F_NAVY)
hdr_cell(wr, 3, 1, 'Metric', F_NAVY)
wr.column_dimensions['A'].width = 32

# Year headers
year_groups = [(1,12,'Year 1',F_BLUE),(13,24,'Year 2',F_TEAL),(25,36,'Year 3',F_GREEN)]
for start,end,label,col in year_groups:
    wr.merge_cells(start_row=2, start_column=start+1, end_row=2, end_column=end+1)
    c = wr.cell(row=2, column=start+1, value=label)
    c.font = font(bold=True, color=F_WHITE, size=11); c.fill = fill(col)
    c.alignment = align('center','center')

for i, m in enumerate(months):
    hdr_cell(wr, 3, i+2, m, F_DKGREY)
    wr.column_dimensions[get_column_letter(i+2)].width = 9

# Revenue data with formulas
# Growth rates per tier per year
home_new  = [1,1,2,2,2,3,3,3,3,4,4,4,  6,6,7,7,8,8,9,9,10,10,10,10,  14,15,16,17,18,19,20,21,22,22,22,22]
care_new  = [0,1,1,1,1,1,1,2,2,2,2,2,  3,4,4,5,5,5,6,6,6,7,7,7,       9,10,11,12,12,13,13,14,14,14,14,14]
pro_new   = [0,0,0,0,1,0,0,1,0,0,1,0,  1,1,1,1,2,1,2,1,2,1,2,1,       3,3,4,4,4,4,5,5,4,4,4,4]
home_churn, care_churn, pro_churn = 0.05, 0.03, 0.02

rows_data = []
h_active = [0]*36; c_active=[0]*36; p_active=[0]*36
for i in range(36):
    if i==0:
        h_active[i] = home_new[i]
        c_active[i] = care_new[i]
        p_active[i] = pro_new[i]
    else:
        h_active[i] = round(h_active[i-1]*(1-home_churn) + home_new[i])
        c_active[i] = round(c_active[i-1]*(1-care_churn) + care_new[i])
        p_active[i] = round(p_active[i-1]*(1-pro_churn)  + pro_new[i])

home_mrr = [h*49 for h in h_active]
care_mrr = [c*199 for c in c_active]
pro_mrr  = [p*499 for p in p_active]
total_mrr = [a+b+c for a,b,c in zip(home_mrr,care_mrr,pro_mrr)]
hw_rev = [round((home_new[i]+care_new[i]+pro_new[i])*0.8*299) for i in range(36)]
op_rev_y = [1, 2, 2, 2, 2, 2, 3, 3, 3, 3, 3, 3,
            4, 5, 5, 6, 6, 6, 6, 6, 7, 7, 7, 7,
            9,10,10,11,12,12,14,14,15,15,15,15]
op_rev = [round(x*416) for x in op_rev_y]
total_rev = [m+h+o for m,h,o in zip(total_mrr,hw_rev,op_rev)]
cum_arr = total_mrr

metrics = [
    ('NEW CUSTOMERS', None),
    ('   New Home Customers', home_new),
    ('   New Care Customers', care_new),
    ('   New Pro Customers', pro_new),
    ('ACTIVE CUSTOMERS', None),
    ('   Active Home', h_active),
    ('   Active Care', c_active),
    ('   Active Pro', p_active),
    ('   TOTAL ACTIVE', [a+b+c for a,b,c in zip(h_active,c_active,p_active)]),
    ('MRR BY TIER', None),
    ('   Home MRR (£)', home_mrr),
    ('   Care MRR (£)', care_mrr),
    ('   Pro MRR (£)', pro_mrr),
    ('   TOTAL MRR (£)', total_mrr),
    ('OTHER REVENUE', None),
    ('   Hardware Kits (£)', hw_rev),
    ('   On-Premise / Support (£)', op_rev),
    ('TOTAL MONTHLY REVENUE (£)', total_rev),
    ('ANNUALISED ARR (£)', [m*12 for m in total_mrr]),
]

fmt_map = {
    'Home MRR (£)': '£#,##0', 'Care MRR (£)': '£#,##0', 'Pro MRR (£)': '£#,##0',
    'TOTAL MRR (£)': '£#,##0', 'Hardware Kits (£)': '£#,##0',
    'On-Premise / Support (£)': '£#,##0', 'TOTAL MONTHLY REVENUE (£)': '£#,##0',
    'ANNUALISED ARR (£)': '£#,##0',
}

section_bgs = {'NEW CUSTOMERS','ACTIVE CUSTOMERS','MRR BY TIER','OTHER REVENUE'}
total_rows  = {'   TOTAL ACTIVE','   TOTAL MRR (£)','TOTAL MONTHLY REVENUE (£)','ANNUALISED ARR (£)'}

row = 4
for label, vals in metrics:
    is_section = label.strip() in section_bgs
    is_total   = label in total_rows
    bg_lbl = F_DKGREY if is_section else (F_NAVY if is_total else F_LGREY)
    fg_lbl = F_WHITE if (is_section or is_total) else F_DKGREY

    c = wr.cell(row=row, column=1, value=label)
    c.fill = fill(bg_lbl); c.font = font(bold=(is_section or is_total), color=fg_lbl)
    c.alignment = align('left','center')
    c.border = border()

    if vals:
        fmt = None
        for k,v in fmt_map.items():
            if k in label: fmt = v; break
        for i, val in enumerate(vals):
            bg = F_LGREY if (i//12)%2==0 else F_WHITE
            if is_total: bg = 'E8F5E9'
            data_cell(wr, row, i+2, val, fmt=fmt, bg=bg,
                      bold=is_total, color=(F_GREEN if is_total else '000000'))
    wr.row_dimensions[row].height = 18
    row += 1

# ── P&L Sheet ─────────────────────────────────────────────────────────────────
wp = wb.create_sheet('P&L')
wp.sheet_view.showGridLines = False
wp.freeze_panes = 'C4'

wp.merge_cells('A1:AK1')
c = wp['A1']; c.value = f'{TRADING_AS} — Profit & Loss (36 Months)'
c.font = font(bold=True, color=F_WHITE, size=13); c.fill = fill(F_NAVY)
c.alignment = align('center','center'); wp.row_dimensions[1].height = 28

hdr_cell(wp, 2, 1, 'Year', F_NAVY); hdr_cell(wp, 3, 1, 'Metric', F_NAVY)
wp.column_dimensions['A'].width = 35

for start,end,label,col in year_groups:
    wp.merge_cells(start_row=2, start_column=start+1, end_row=2, end_column=end+1)
    c = wp.cell(row=2, column=start+1, value=label)
    c.font=font(bold=True,color=F_WHITE,size=11); c.fill=fill(col)
    c.alignment=align('center','center')

for i,m in enumerate(months):
    hdr_cell(wp, 3, i+2, m, F_DKGREY)
    wp.column_dimensions[get_column_letter(i+2)].width = 9

# Cost model
n_cust = [a+b+c for a,b,c in zip(h_active,c_active,p_active)]
hosting = [150 + (n//100)*40 for n in n_cust]
sms_email = [50 + n*0.3 for n in n_cust]
tools = [40]*36
stripe = [round(total_rev[i]*0.015) for i in range(36)]
mktg = [0]*12 + [1500]*12 + [4000]*12
total_opex = [h+s+t+st+m for h,s,t,st,m in zip(hosting,sms_email,tools,stripe,mktg)]
hw_cogs = [round(hw_rev[i]*0.55) for i in range(36)]
total_cogs = [h+o for h,o in zip(hw_cogs,[round(x*0.3) for x in op_rev])]
gross_profit = [total_rev[i]-total_cogs[i] for i in range(36)]
ebitda = [gross_profit[i]-total_opex[i] for i in range(36)]
cum_ebitda = []
running=0
for e in ebitda:
    running+=e; cum_ebitda.append(running)

pl_metrics = [
    ('REVENUE', None, None),
    ('   SaaS / MRR Revenue (£)', total_mrr, '£#,##0'),
    ('   Hardware Revenue (£)', hw_rev, '£#,##0'),
    ('   On-Premise / Support (£)', op_rev, '£#,##0'),
    ('TOTAL REVENUE (£)', total_rev, '£#,##0'),
    ('COST OF GOODS SOLD', None, None),
    ('   Hardware COGS (£)', hw_cogs, '£#,##0'),
    ('   On-Premise Delivery (£)', [round(x*0.3) for x in op_rev], '£#,##0'),
    ('TOTAL COGS (£)', total_cogs, '£#,##0'),
    ('GROSS PROFIT (£)', gross_profit, '£#,##0'),
    ('GROSS MARGIN (%)', [round(gross_profit[i]/total_rev[i]*100,1) if total_rev[i]>0 else 0 for i in range(36)], '0.0"%"'),
    ('OPERATING EXPENSES', None, None),
    ('   Cloud Hosting (£)', hosting, '£#,##0'),
    ('   Email + SMS (£)', [round(x) for x in sms_email], '£#,##0'),
    ('   Software Tools (£)', tools, '£#,##0'),
    ('   Stripe / Payment Fees (£)', stripe, '£#,##0'),
    ('   Marketing (£)', mktg, '£#,##0'),
    ('TOTAL OPEX (£)', [round(x) for x in total_opex], '£#,##0'),
    ('EBITDA (£)', ebitda, '£#,##0'),
    ('CUMULATIVE EBITDA (£)', cum_ebitda, '£#,##0'),
]

section_bg_pl = {'REVENUE','COST OF GOODS SOLD','OPERATING EXPENSES'}
total_rows_pl = {'TOTAL REVENUE (£)','TOTAL COGS (£)','GROSS PROFIT (£)',
                 'GROSS MARGIN (%)','TOTAL OPEX (£)','EBITDA (£)','CUMULATIVE EBITDA (£)'}

row=4
for label, vals, fmt in pl_metrics:
    is_sec = label.strip() in section_bg_pl
    is_tot = label in total_rows_pl
    is_ebitda = 'EBITDA' in label
    bg_l = F_DKGREY if is_sec else (F_NAVY if is_tot else F_LGREY)
    fg_l = F_WHITE if (is_sec or is_tot) else F_DKGREY

    c = wp.cell(row=row, column=1, value=label)
    c.fill=fill(bg_l); c.font=font(bold=(is_sec or is_tot), color=fg_l)
    c.alignment=align('left','center'); c.border=border()

    if vals:
        for i,val in enumerate(vals):
            bg = 'E8F5E9' if is_tot else (F_LGREY if (i//12)%2==0 else F_WHITE)
            color = F_RED if (is_ebitda and isinstance(val,(int,float)) and val<0) else \
                    (F_GREEN if is_tot else '000000')
            data_cell(wp, row, i+2, val, fmt=fmt, bg=bg,
                      bold=is_tot, color=color)
    wp.row_dimensions[row].height = 18
    row+=1

# ── Break-Even Sheet ──────────────────────────────────────────────────────────
wb_ = wb.create_sheet('Break-Even')
wb_.sheet_view.showGridLines = False
set_col_widths(wb_, [3, 30, 18, 18, 18, 30])

wb_.merge_cells('B1:F1')
c = wb_['B1']; c.value = f'{TRADING_AS} — Break-Even Analysis'
c.font=font(bold=True,color=F_WHITE,size=13); c.fill=fill(F_NAVY)
c.alignment=align('center','center'); wb_.row_dimensions[1].height=30

for col,h_ in enumerate(['Scenario','Fixed Costs/Mo','Revenue/Customer','Customers Needed','Monthly Profit at 3×'],2):
    hdr_cell(wb_, 2, col, h_, F_BLUE)

be_data = [
    ('Absolute minimum\n(1 tier: Home only)', 280, 49, None, None),
    ('Realistic minimum\n(mix of tiers)', 280, 120, None, None),
    ('With Year 1 marketing\n(£0 paid ads)', 320, 120, None, None),
    ('With Year 2 marketing\n(£1,500/mo ads)', 1820, 150, None, None),
    ('With Year 3 marketing\n(£4,000/mo ads)', 4320, 175, None, None),
]

row=3
for scenario, fixed, rev_per, _, __ in be_data:
    needed = round(fixed / rev_per + 0.5)
    profit_3x = needed*3*rev_per - fixed
    bg = F_LGREY if row%2==0 else F_WHITE
    label_cell(wb_, row, 2, scenario, bg=bg)
    data_cell(wb_, row, 3, fixed, '£#,##0', bg=bg)
    data_cell(wb_, row, 4, rev_per, '£#,##0', bg=bg)
    c = wb_.cell(row=row, column=5, value=needed)
    c.fill=fill('FFF9C4'); c.font=font(bold=True,color=F_NAVY)
    c.alignment=align('center','center'); c.border=border()
    data_cell(wb_, row, 6, profit_3x, '£#,##0', bg='E8F5E9', color=F_GREEN, bold=True)
    row+=1

wb_.cell(row=row+1, column=2,
    value='* Fixed costs include hosting, email/SMS, tools. Excludes founder salary.').font = \
    font(italic=True, color=F_MGREY, size=9)
wb_.cell(row=row+2, column=2,
    value='* Revenue per customer is blended average across tiers.').font = \
    font(italic=True, color=F_MGREY, size=9)

# ── Cash Flow Sheet ───────────────────────────────────────────────────────────
wc = wb.create_sheet('Cash Flow')
wc.sheet_view.showGridLines = False
wc.freeze_panes = 'C4'
wc.column_dimensions['A'].width = 35

wc.merge_cells('A1:AK1')
c = wc['A1']; c.value = f'{TRADING_AS} — Cash Flow (36 Months)'
c.font=font(bold=True,color=F_WHITE,size=13); c.fill=fill(F_NAVY)
c.alignment=align('center','center'); wc.row_dimensions[1].height=28

hdr_cell(wc,2,1,'Year',F_NAVY); hdr_cell(wc,3,1,'Item',F_NAVY)
for start,end,label,col in year_groups:
    wc.merge_cells(start_row=2,start_column=start+1,end_row=2,end_column=end+1)
    c=wc.cell(row=2,column=start+1,value=label)
    c.font=font(bold=True,color=F_WHITE,size=11); c.fill=fill(col)
    c.alignment=align('center','center')
for i,m in enumerate(months):
    hdr_cell(wc,3,i+2,m,F_DKGREY)
    wc.column_dimensions[get_column_letter(i+2)].width=9

# Opening balance — assume starting with £0
opening = [0]+[None]*35
total_in  = [total_rev[i] for i in range(36)]
hw_out    = [round(hw_rev[i]*0.55) for i in range(36)]  # buying hardware stock
total_out = [round(total_opex[i]+hw_out[i]) for i in range(36)]
net_cf    = [total_in[i]-total_out[i] for i in range(36)]
closing   = []
bal=0
for n in net_cf:
    bal+=n; closing.append(bal)
opening_list = [0]+closing[:-1]

cf_rows = [
    ('CASH IN', None, None),
    ('   Subscription Revenue (£)', total_mrr, '£#,##0'),
    ('   Hardware Sales (£)', hw_rev, '£#,##0'),
    ('   Other Revenue (£)', op_rev, '£#,##0'),
    ('TOTAL CASH IN (£)', total_in, '£#,##0'),
    ('CASH OUT', None, None),
    ('   Hardware Stock Purchase (£)', hw_out, '£#,##0'),
    ('   Operating Expenses (£)', [round(x) for x in total_opex], '£#,##0'),
    ('TOTAL CASH OUT (£)', total_out, '£#,##0'),
    ('NET CASH FLOW (£)', net_cf, '£#,##0'),
    ('OPENING BALANCE (£)', opening_list, '£#,##0'),
    ('CLOSING BALANCE (£)', closing, '£#,##0'),
]

row=4
cf_totals = {'TOTAL CASH IN (£)','TOTAL CASH OUT (£)','NET CASH FLOW (£)',
             'OPENING BALANCE (£)','CLOSING BALANCE (£)'}
for label,vals,fmt in cf_rows:
    is_sec = label in {'CASH IN','CASH OUT'}
    is_tot = label in cf_totals
    bg_l=F_DKGREY if is_sec else (F_NAVY if is_tot else F_LGREY)
    fg_l=F_WHITE if (is_sec or is_tot) else F_DKGREY
    c=wc.cell(row=row,column=1,value=label)
    c.fill=fill(bg_l); c.font=font(bold=(is_sec or is_tot),color=fg_l)
    c.alignment=align('left','center'); c.border=border()
    if vals:
        for i,val in enumerate(vals):
            bg='E8F5E9' if 'CLOSING' in label else (F_LGREY if (i//12)%2==0 else F_WHITE)
            color=F_RED if (isinstance(val,(int,float)) and val<0 and 'NET' in label) else \
                  (F_GREEN if is_tot else '000000')
            data_cell(wc,row,i+2,val,fmt=fmt,bg=bg,bold=is_tot,color=color)
    wc.row_dimensions[row].height=18; row+=1

# Add revenue chart to P&L sheet
chart = LineChart()
chart.title = 'Monthly Revenue & EBITDA'
chart.style = 10
chart.y_axis.title = '£'
chart.x_axis.title = 'Month'
chart.height = 12; chart.width = 22

rev_ref = Reference(wp, min_col=2, max_col=37,
                     min_row=[r for r,_ in enumerate(pl_metrics,4) if _[0]=='TOTAL REVENUE (£)'][0] if any(_[0]=='TOTAL REVENUE (£)' for _ in pl_metrics) else 6,
                     max_row=[r for r,_ in enumerate(pl_metrics,4) if _[0]=='TOTAL REVENUE (£)'][0] if any(_[0]=='TOTAL REVENUE (£)' for _ in pl_metrics) else 6)

# Save
path = os.path.join(OUT, 'EchoPose_Financial_Model.xlsx')
wb.save(path); print(f'  Saved: {path}')


# ══════════════════════════════════════════════════════════════════════════════
#  CAP TABLE
# ══════════════════════════════════════════════════════════════════════════════
print('Building Cap Table...')
ct = Workbook()
ws = ct.active; ws.title = 'Cap Table'
ws.sheet_view.showGridLines = False

ws.merge_cells('A1:K1')
c = ws['A1']; c.value = f'{TRADING_AS} — Capitalisation Table'
c.font=font(bold=True,color=F_WHITE,size=14); c.fill=fill(F_NAVY)
c.alignment=align('center','center'); ws.row_dimensions[1].height=32

ws.merge_cells('A2:K2')
c=ws['A2']; c.value=f'Prepared by: {OWNER}  |  Status: Pre-Company / Sole Trader  |  April 2026'
c.font=font(color=F_MGREY,size=9,italic=True); c.fill=fill(F_DKGREY)
c.alignment=align('center','center'); ws.row_dimensions[2].height=18

# ── Section 1: Current state ──────────────────────────────────────────────────
ws.merge_cells('A4:K4')
c=ws['A4']; c.value='CURRENT STATE — PRE-INCORPORATION'
c.font=font(bold=True,color=F_WHITE,size=11); c.fill=fill(F_ORANGE)
c.alignment=align('left','center'); ws.row_dimensions[4].height=22

ws['A5'] = 'As a sole trader, there are no shares. When you incorporate as a Ltd company, '
ws['A5'].font = font(italic=True, color=F_DKGREY)
ws.merge_cells('A5:K5')
ws['A6'] = 'the cap table becomes active. This sheet shows what it will look like at each stage.'
ws['A6'].font = font(italic=True, color=F_DKGREY)
ws.merge_cells('A6:K6')

# ── Section 2: Post-incorporation ─────────────────────────────────────────────
ws.merge_cells('A8:K8')
c=ws['A8']; c.value='STAGE 1 — POST-INCORPORATION (Day 1 as Ltd Company)'
c.font=font(bold=True,color=F_WHITE,size=11); c.fill=fill(F_BLUE)
c.alignment=align('left','center'); ws.row_dimensions[8].height=22

hdrs = ['Shareholder','Role','Share Class','Shares Held','% Ownership','Issue Price (£)','Total Value (£)','Notes']
col_w = [30,20,15,14,14,14,14,35]
for i,(h_,w) in enumerate(zip(hdrs,col_w),1):
    hdr_cell(ws, 9, i, h_, F_DKGREY)
    ws.column_dimensions[get_column_letter(i)].width = w

s1_rows = [
    [OWNER, 'Founder / Owner', 'Ordinary', 1000000, '100.0%', 0.001, 1000, 'Founder shares at par value'],
    ['Employee Option Pool', '(Reserved)', 'EMI Options', 0, '0.0%', '—', '—', 'Reserve 10–15% for future hires. Issue via EMI scheme.'],
    ['TOTAL', '', '', 1000000, '100.0%', '', 1000, ''],
]
for ri, row_ in enumerate(s1_rows):
    bg = F_LGREY if ri%2==0 else F_WHITE
    is_total = row_[0]=='TOTAL'
    for ci,val in enumerate(row_,1):
        c=ws.cell(row=10+ri, column=ci, value=val)
        c.fill=fill(F_NAVY if is_total else bg)
        c.font=font(bold=is_total, color=F_WHITE if is_total else '000000')
        c.alignment=align('center','center') if ci>2 else align('left','center')
        c.border=border()
        if ci==5 and not is_total: c.number_format='0.0%'
        if ci in (6,7) and isinstance(val,(int,float)):
            c.number_format='£#,##0.000' if ci==6 else '£#,##0'

# ── Section 3: First investment round ─────────────────────────────────────────
ws.merge_cells('A15:K15')
c=ws['A15']; c.value='STAGE 2 — SEED ROUND (Hypothetical — if you raise external investment)'
c.font=font(bold=True,color=F_WHITE,size=11); c.fill=fill(F_TEAL)
c.alignment=align('left','center'); ws.row_dimensions[15].height=22

for i,h_ in enumerate(hdrs,1):
    hdr_cell(ws, 16, i, h_, F_DKGREY)

seed_invest = 50000  # £50k seed
pre_money   = 500000 # £500k pre-money valuation
new_shares  = round(1000000 * seed_invest / pre_money)
total_post  = 1000000 + new_shares
founder_pct = 1000000/total_post
investor_pct= new_shares/total_post

s2_rows = [
    [OWNER, 'Founder / Owner', 'Ordinary', 1000000, f'{founder_pct*100:.1f}%', 0.001, round(founder_pct*pre_money+seed_invest), 'Post-round valuation on founder shares'],
    ['Angel Investor / VC', 'Investor', 'Preferred', new_shares, f'{investor_pct*100:.1f}%', round(seed_invest/new_shares,3), seed_invest, f'£{seed_invest:,} for {investor_pct*100:.1f}% at £{pre_money:,} pre-money'],
    ['Option Pool (10%)', 'Employees', 'EMI Options', round(total_post*0.1), '~10.0%', '—', '—', 'Authorised but unissued'],
    ['TOTAL (post-round)','','', round(total_post*1.1), '100.0%', '', round(pre_money+seed_invest), ''],
]
for ri,row_ in enumerate(s2_rows):
    bg = F_LGREY if ri%2==0 else F_WHITE
    is_total = 'TOTAL' in str(row_[0])
    for ci,val in enumerate(row_,1):
        c=ws.cell(row=17+ri, column=ci, value=val)
        c.fill=fill(F_NAVY if is_total else bg)
        c.font=font(bold=is_total, color=F_WHITE if is_total else '000000')
        c.alignment=align('center','center') if ci>2 else align('left','center')
        c.border=border()

# ── Section 4: Key concepts ────────────────────────────────────────────────────
ws.merge_cells('A23:K23')
c=ws['A23']; c.value='KEY CONCEPTS — READ THIS BEFORE INCORPORATING'
c.font=font(bold=True,color=F_WHITE,size=11); c.fill=fill(F_NAVY)
c.alignment=align('left','center'); ws.row_dimensions[23].height=22

concepts = [
    ('Pre-Money Valuation', 'What your company is worth BEFORE investment. You set this in negotiation. Higher = less dilution.'),
    ('Post-Money Valuation', 'Pre-money + the investment received. E.g. £500k pre + £50k investment = £550k post-money.'),
    ('Dilution', 'When new shares are issued (for investment or options), your % ownership decreases. This is normal and expected.'),
    ('EMI Option Scheme', 'UK government scheme to give employees share options with tax advantages. Highly recommended for first hires.'),
    ('Ordinary Shares', 'Standard shares with voting rights. You hold these as founder.'),
    ('Preferred Shares', 'Investor shares — often have preference on liquidation (they get paid first if company is sold).'),
    ('Option Pool', 'Shares reserved for future employees. Typically 10–15% created before raising investment.'),
    ('SEIS / EIS', 'UK tax relief schemes for investors in early-stage companies. Makes investing in you more attractive to angels.'),
]
for ri,(term,desc) in enumerate(concepts):
    ws.cell(row=24+ri,column=1,value=term).font=font(bold=True,color=F_BLUE)
    c=ws.cell(row=24+ri,column=2,value=desc)
    c.font=font(color=F_DKGREY); c.alignment=align('left','center',wrap=True)
    ws.merge_cells(start_row=24+ri,start_column=2,end_row=24+ri,end_column=8)
    ws.row_dimensions[24+ri].height=30

path = os.path.join(OUT, 'EchoPose_Cap_Table.xlsx')
ct.save(path); print(f'  Saved: {path}')


# ══════════════════════════════════════════════════════════════════════════════
#  INVOICE / QUOTE TEMPLATE
# ══════════════════════════════════════════════════════════════════════════════
print('Building Invoice & Quote Templates...')
iv = Workbook()

for sheet_name, doc_type, accent in [
    ('Invoice Template', 'INVOICE', F_NAVY),
    ('Quote Template', 'QUOTE / PROPOSAL', F_BLUE),
]:
    if sheet_name == 'Invoice Template':
        ws = iv.active; ws.title = sheet_name
    else:
        ws = iv.create_sheet(sheet_name)

    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    for col, w in enumerate([2,18,28,12,12,12,14,2], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Header block
    ws.merge_cells('B2:D6')
    c=ws['B2']
    c.value = f'{TRADING_AS}\n{OWNER}\n{EMAIL}\nUnited Kingdom'
    c.font=font(bold=False, color=F_WHITE, size=11); c.fill=fill(F_NAVY)
    c.alignment=Alignment(horizontal='left', vertical='top',
                          wrap_text=True, indent=1)

    ws.merge_cells('E2:G2')
    c=ws['E2']; c.value=doc_type
    c.font=font(bold=True, color=F_WHITE, size=22); c.fill=fill(accent)
    c.alignment=align('right','center')
    ws.row_dimensions[2].height=36

    for r,h_ in [(3,'NUMBER:'),(4,'DATE:'),(5,'DUE DATE:'),(6,'STATUS:')]:
        ws.cell(row=r,column=5,value=h_).font=font(bold=True,color=F_WHITE); \
        ws.cell(row=r,column=5).fill=fill(F_DKGREY)
        ws.cell(row=r,column=5).alignment=align('right','center')
        placeholder = {'NUMBER:':'INV-001','DATE:':'18/04/2026',
                       'DUE DATE:':'02/05/2026','STATUS:':'UNPAID'}.get(h_,'')
        if doc_type=='QUOTE / PROPOSAL':
            placeholder = {'NUMBER:':'QUO-001','DATE:':'18/04/2026',
                           'DUE DATE:':'Valid 30 days','STATUS:':'DRAFT'}.get(h_,placeholder)
        ws.merge_cells(start_row=r,start_column=6,end_row=r,end_column=7)
        c=ws.cell(row=r,column=6,value=placeholder)
        c.font=font(bold=True,color=F_NAVY); c.fill=fill('FFFDE7')
        c.alignment=align('right','center'); c.border=border()
        ws.row_dimensions[r].height=20

    # Bill to
    ws.cell(row=8,column=2,value='BILL TO:').font=font(bold=True,color=F_WHITE); \
    ws.cell(row=8,column=2).fill=fill(accent)
    ws.cell(row=8,column=2).alignment=align('left','center')
    for r,ph in [(9,'Customer / Organisation Name'),(10,'Contact Name'),
                 (11,'Address Line 1'),(12,'Address Line 2'),(13,'Email')]:
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
        c=ws.cell(row=r,column=2,value=ph)
        c.font=font(color=F_MGREY); c.fill=fill(F_LGREY)
        c.alignment=align('left','center'); c.border=border()
        ws.row_dimensions[r].height=18

    ws.row_dimensions[8].height=20; ws.row_dimensions[14].height=8

    # Line items header
    for col,txt,bg in [(2,'#',F_NAVY),(3,'Description',F_NAVY),(4,'Qty',F_NAVY),
                        (5,'Unit',F_NAVY),(6,'Unit Price (£)',F_NAVY),(7,'Total (£)',F_NAVY)]:
        c=ws.cell(row=15,column=col,value=txt)
        c.font=font(bold=True,color=F_WHITE); c.fill=fill(bg)
        c.alignment=align('center','center'); c.border=border()
    ws.row_dimensions[15].height=20

    # Line item rows
    sample_items = [
        ('1','EchoPose Care — Monthly Subscription','1','month','199.00','=D17*F17'),
        ('2','Hardware Starter Kit (3× ESP32-S3 + TP-Link AP)','1','kit','299.00','=D18*F18'),
        ('3','Professional Setup & Calibration (remote)','1','session','150.00','=D19*F19'),
        ('4','','','','',''),
        ('5','','','','',''),
    ]
    for ri,(num,desc,qty,unit,price,total) in enumerate(sample_items):
        r=16+ri; bg=F_LGREY if ri%2==0 else F_WHITE
        for col,val in [(2,num),(3,desc),(4,qty if qty else ''),(5,unit),(6,float(price) if price else ''),(7,total)]:
            c=ws.cell(row=r,column=col,value=val if val!='' else '')
            c.fill=fill(bg)
            c.font=font(color='000000')
            c.alignment=align('center','center') if col!=3 else align('left','center')
            c.border=border()
            if col==6 and isinstance(val,float): c.number_format='£#,##0.00'
            if col==7 and val and val!='': c.number_format='£#,##0.00'
        ws.row_dimensions[r].height=22

    # Totals block
    subtotal_row = 22
    ws.merge_cells(f'B{subtotal_row}:E{subtotal_row}')
    ws.cell(row=subtotal_row,column=2,value='SUBTOTAL').font=font(bold=True,color=F_DKGREY)
    ws.cell(row=subtotal_row,column=2).fill=fill(F_LGREY)
    ws.cell(row=subtotal_row,column=2).alignment=align('right','center')
    c=ws.cell(row=subtotal_row,column=6)
    c.value=f'=SUM(G16:G{subtotal_row-1})'
    c.number_format='£#,##0.00'; c.font=font(bold=True)
    c.fill=fill(F_LGREY); c.border=border()
    ws.merge_cells(f'F{subtotal_row}:G{subtotal_row}')

    vat_row=subtotal_row+1
    ws.merge_cells(f'B{vat_row}:E{vat_row}')
    ws.cell(row=vat_row,column=2,value='VAT (0% — Not VAT Registered)').font=font(italic=True,color=F_MGREY)
    ws.cell(row=vat_row,column=2).fill=fill(F_LGREY)
    ws.cell(row=vat_row,column=2).alignment=align('right','center')
    c=ws.cell(row=vat_row,column=6,value=0.00)
    c.number_format='£#,##0.00'; c.fill=fill(F_LGREY); c.border=border()
    ws.merge_cells(f'F{vat_row}:G{vat_row}')

    total_row=vat_row+1
    ws.merge_cells(f'B{total_row}:E{total_row}')
    c=ws.cell(row=total_row,column=2,value='TOTAL DUE')
    c.font=font(bold=True,color=F_WHITE,size=13); c.fill=fill(accent)
    c.alignment=align('right','center')
    c=ws.cell(row=total_row,column=6,value=f'=G{subtotal_row}+G{vat_row}')
    c.number_format='£#,##0.00'; c.font=font(bold=True,color=F_WHITE,size=13)
    c.fill=fill(accent); c.border=border()
    ws.merge_cells(f'F{total_row}:G{total_row}')
    ws.row_dimensions[total_row].height=28

    # Payment details
    pay_row=total_row+2
    ws.merge_cells(f'B{pay_row}:G{pay_row}')
    c=ws.cell(row=pay_row,column=2,value='PAYMENT DETAILS')
    c.font=font(bold=True,color=F_WHITE); c.fill=fill(F_DKGREY)
    c.alignment=align('left','center')

    pay_items = [
        ('Bank / Transfer', 'Monzo / Revolut — details provided separately'),
        ('Reference', 'Please use invoice number as payment reference'),
        ('PayPal / Stripe', 'Payment link will be sent via email if preferred'),
        ('Payment Terms', '14 days from invoice date' if doc_type=='INVOICE' else 'Quote valid for 30 days from issue date'),
    ]
    for ri,(k,v) in enumerate(pay_items):
        r=pay_row+1+ri; bg=F_LGREY if ri%2==0 else F_WHITE
        ws.cell(row=r,column=2,value=k).font=font(bold=True,color=F_NAVY)
        ws.cell(row=r,column=2).fill=fill(bg); ws.cell(row=r,column=2).border=border()
        ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=7)
        c=ws.cell(row=r,column=3,value=v)
        c.fill=fill(bg); c.border=border(); ws.row_dimensions[r].height=18

    # Footer
    fr = pay_row+6
    ws.merge_cells(f'B{fr}:G{fr}')
    c=ws.cell(row=fr,column=2,
        value=f'{TRADING_AS}  ·  {OWNER}  ·  {EMAIL}  ·  Not VAT Registered')
    c.font=font(color=F_MGREY,size=8,italic=True)
    c.alignment=align('center','center')

path = os.path.join(OUT, 'EchoPose_Invoice_Quote_Template.xlsx')
iv.save(path); print(f'  Saved: {path}')

print(f'\nAll financial documents saved to {OUT}/')
